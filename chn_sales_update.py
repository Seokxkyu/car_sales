import os
import argparse
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
EXCEL_FILE = os.path.join(DATA_DIR, 'china_sales_update.xlsx')

brand_map = {
    '比亚迪':'BYD','大众':'VW','吉利':'Geely','丰田':'Toyota','奇瑞':'Chery',
    '长安':'Changan','本田':'Honda','特斯拉':'Tesla','奥迪':'Audi','五菱':'Wuling',
    '五菱（银标）':'Wuling (Silver)','捷途':'Jetour','奔驰':'Benz','哈弗':'Haval',
    'MG':'MG','宝马':'BMW','银河':'GALAXY','日产':'Nissan','红旗':'Hongqi',
    '零跑':'Leapmotor','理想':'LI','别克':'Buick','小鹏':'XPENG','传祺':'GAC Trumpchi',
    '小米':'XIAOMI','领克':'Lynk&CO','埃安':'Aion','起亚':'Kia','荣威':'Roewe',
    '坦克':'TANK','现代':'Hyundai'
}

def fetch_china_sales(ym, url):
    resp = requests.get(url, timeout=10)
    resp.encoding = resp.apparent_encoding
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, 'lxml')
    table = soup.find('table')
    if not table:
        raise RuntimeError(f"{ym} 페이지에서 테이블을 찾을 수 없습니다")
    records = []
    for tr in table.find_all('tr')[1:]:
        tds = tr.find_all('td')
        if len(tds) < 3:
            continue
        cn = tds[1].get_text(strip=True)
        txt = tds[2].get_text(strip=True).replace(',', '')
        try:
            sales = int(txt)
        except ValueError:
            continue
        en = brand_map.get(cn)
        if en:
            records.append({'Brand': en, 'Sales': sales})
    if not records:
        raise RuntimeError(f"{ym} 데이터가 없습니다")
    df = pd.DataFrame(records).set_index('Brand')
    month_date = pd.to_datetime(f"{ym}-01").normalize()
    df.columns = [month_date]
    return df

def update_china_sales(ym, url, sheet_name='Brands'):
    new_df = fetch_china_sales(ym, url)
    date = new_df.columns[0]
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    # 첫 열 Brand 헤더
    if ws.cell(row=1, column=1).value is None:
        ws.cell(row=1, column=1, value='Brand')
    # 새 컬럼 인덱스 및 헤더 설정
    new_col = ws.max_column + 1
    header_cell = ws.cell(row=1, column=new_col, value=date.strftime('%Y-%m'))
    header_cell.alignment = Alignment(horizontal='center')
    header_cell.font = Font(bold=True)
    # 기존 브랜드 매핑
    brand_rows = { ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row+1) }
    # 데이터 입력 및 숫자 포맷
    for brand, sales in new_df[date].items():
        row = brand_rows.get(brand, ws.max_row+1)
        if brand not in brand_rows:
            ws.cell(row=row, column=1, value=brand)
        cell = ws.cell(row=row, column=new_col, value=sales)
        cell.number_format = '#,###,###'
    ws.freeze_panes = 'B2'
    wb.save(EXCEL_FILE)
    print(f"'{EXCEL_FILE}' 업데이트 완료")

def main():
    parser = argparse.ArgumentParser(prog='chn_sales_update', description='중국 자동차 판매 업데이트')
    parser.add_argument('year_month', help='YYYY-MM')
    parser.add_argument('url', help='스크래핑 URL')
    args = parser.parse_args()
    try:
        update_china_sales(args.year_month, args.url)
    except Exception as e:
        print(f"업데이트 실패: {e}")

if __name__ == '__main__':
    main()
