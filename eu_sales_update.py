import os
import calendar
import requests
import pandas as pd
import camelot
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import argparse

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_DIR = os.path.join(BASE_DIR, 'acea_pdfs')
DATA_DIR = os.path.join(BASE_DIR, 'data')

os.makedirs(PDF_DIR, exist_ok=True)
HEADERS = {"User-Agent": "Mozilla/5.0"}

def parse_europe_page6(pdf_path: str) -> pd.DataFrame:
    tables = camelot.read_pdf(
        pdf_path,
        pages='6',
        flavor='stream',
        strip_text='\n'
    )
    if not tables:
        raise RuntimeError(f"테이블 파싱 실패: {pdf_path}")
    df = tables[0].df
    header1 = df.iloc[2]
    header2 = df.iloc[3]
    data = df.iloc[4:].reset_index(drop=True)
    data.columns = pd.MultiIndex.from_arrays([header1, header2])
    result = data.iloc[:, [0, 3]].reset_index(drop=True)
    result.columns = [' '.join(map(str, col)).strip() for col in result.columns.values]
    result.rename(columns={result.columns[0]: "Manufacturer", result.columns[1]: "value"}, inplace=True)
    result["Manufacturer"] = result["Manufacturer"].str.replace(r"\d+", "", regex=True).str.strip()
    result['value'] = result['value'].astype(str).str.replace(',', '').astype(int)
    return result

def update_excel_with_new_column(
    excel_file: str,
    year_month: str,
    pdf_file: str,
    sheet_name="Europe"
):
    new_df = parse_europe_page6(pdf_file)

    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    df_old = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl", dtype=str)

    if year_month in df_old.columns:
        print(f"[Skip] {year_month} 컬럼이 이미 있습니다!")
        return

    new_df['Manufacturer'] = new_df['Manufacturer'].astype(str).str.strip()
    df_old['Manufacturer'] = df_old['Manufacturer'].astype(str).str.strip()

    merged = pd.merge(
        df_old,
        new_df[['Manufacturer', 'value']].rename(columns={'value': year_month}),
        how='left',
        on='Manufacturer'
    )

    col_idx = ws.max_column + 1
    header_cell = ws.cell(row=1, column=col_idx, value=year_month)
    header_cell.font = Font(size=10, bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    man_to_value = dict(zip(merged['Manufacturer'], merged[year_month]))
    for r in range(2, ws.max_row+1):
        manu = ws.cell(row=r, column=1).value
        key = str(manu).strip() if manu is not None else ""
        val = man_to_value.get(key, "")
        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.font = Font(size=10)
        cell.number_format = '#,###'

    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 10.25
    
    wb.save(excel_file)
    print(f"✅ {sheet_name} 시트에 {year_month} 컬럼 추가 완료!")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('pdf_file', help='파싱할 PDF 파일 경로')
    parser.add_argument('year_month', help='추가할 연월 (예: 2025-05)')
    parser.add_argument('-e', '--excel_file', default=os.path.join(DATA_DIR, "europe_sales_update.xlsx"), help='엑셀 파일 경로')
    parser.add_argument('-s', '--sheet', default='Europe', help='시트 이름 (기본: Europe)')
    args = parser.parse_args()

    update_excel_with_new_column(args.excel_file, args.year_month, args.pdf_file, sheet_name=args.sheet)


#!/usr/bin/env python3

# import os
# import argparse
# import calendar
# import requests
# import pandas as pd
# import tabula
# from openpyxl import load_workbook

# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# PDF_DIR = os.path.join(BASE_DIR, 'acea_pdfs')
# DATA_DIR = os.path.join(BASE_DIR, 'data')

# os.makedirs(PDF_DIR, exist_ok=True)

# HEADERS = {"User-Agent": "Mozilla/5.0"}


# def fetch_europe_pdf(year_month: str, url: str) -> str:
#     year, mon = year_month.split('-')
#     month_name = calendar.month_name[int(mon)]
#     fname = f"{year}_{month_name}.pdf"
#     path = os.path.join(PDF_DIR, fname)

#     if os.path.exists(path):
#         print(f"[Info] 이미 다운로드됨: {fname}")
#         return path

#     resp = requests.get(url, headers=HEADERS, timeout=30)
#     resp.raise_for_status()
#     with open(path, 'wb') as f:
#         f.write(resp.content)
#     print(f"[Downloaded] {fname}")
#     return path


# def parse_europe_page6(pdf_path: str) -> pd.DataFrame:
#     # tabula-py를 이용해 6페이지 첫 테이블 읽기
#     tables = tabula.read_pdf(
#         pdf_path,
#         pages='6',
#         multiple_tables=True,
#         stream=True
#     )
#     if not tables:
#         raise RuntimeError(f"테이블 파싱 실패: {pdf_path}")
#     df = tables[0]
#     df.columns = df.iloc[0]
#     return df.iloc[1:].reset_index(drop=True)


# def update_europe_sales_only(
#     excel_file_arg: str,
#     year_month: str,
#     url: str,
#     sheet_name: str = None
# ):
#     # 엑셀 경로 결정
#     if os.path.isabs(excel_file_arg):
#         excel_path = excel_file_arg
#     else:
#         excel_path = os.path.join(DATA_DIR, excel_file_arg)

#     # PDF 다운로드 및 파싱
#     pdf_file = fetch_europe_pdf(year_month, url)
#     df6 = parse_europe_page6(pdf_file)
#     print(df6.columns.tolist())
#     print(df6.head(3))

#     # 시트명 및 mode 설정
#     target_sheet = sheet_name or year_month
#     mode = 'a' if os.path.exists(excel_path) else 'w'
#     writer_kwargs = {'engine': 'openpyxl', 'mode': mode}
#     if mode == 'a':
#         writer_kwargs['if_sheet_exists'] = 'new'

#     # 전처리: 세 번째 컬럼(split 대상)의 첫 번째 값만 추출
#     col_c = df6.columns[2]
#     new_col = target_sheet
#     df6[new_col] = df6[col_c].astype(str).str.split(' ').str[0]
#     df6.drop(columns=[col_c], inplace=True)

#     # 엑셀에 시트 추가 (헤더 없이 데이터만)
#     with pd.ExcelWriter(excel_path, **writer_kwargs) as writer:
#         df6.to_excel(writer, sheet_name=target_sheet, index=False, header=False)

#     print(f"✅ '{excel_path}' 에 시트 '{target_sheet}' 추가 완료!")


# def main():
#     parser = argparse.ArgumentParser(
#         prog='update_europe_sales',
#         description='ACEA PDF(6페이지) → Excel 신규 시트로 추가 (tabula 사용)'
#     )
#     parser.add_argument(
#         'excel_file',
#         help='data 폴더 내 파일명 또는 절대경로'
#     )
#     parser.add_argument(
#         'year_month',
#         help='추가할 연월 (YYYY-MM)'
#     )
#     parser.add_argument(
#         'url',
#         help='해당 연월의 ACEA PDF URL'
#     )
#     parser.add_argument(
#         '-s', '--sheet',
#         default=None,
#         help='시트 이름 (기본: YYYY-MM)'
#     )
#     args = parser.parse_args()
#     try:
#         update_europe_sales_only(
#             args.excel_file,
#             args.year_month,
#             args.url,
#             sheet_name=args.sheet
#         )
#     except Exception as e:
#         print(f"❌ 업데이트 실패: {e}")

# if __name__ == '__main__':
#     main()