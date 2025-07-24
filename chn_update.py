#!/usr/bin/env python3
import os
from openpyxl import load_workbook
from datetime import datetime

Base_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(Base_DIR, 'data')
EXCEL_FILE = os.path.join(DATA_DIR, 'china_sales_update.xlsx')

# 워크북 로드
wb = load_workbook(EXCEL_FILE)
ws = wb['Brands']

# 1행: 헤더를 문자열 'YYYY-MM' 형식으로 통일
for col in range(2, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    if isinstance(cell.value, datetime):
        cell.value = cell.value.strftime('%Y-%m')

# 2행 이하, 2열 이상: 숫자 포맷 '#,###,###' 적용
for row in range(2, ws.max_row + 1):
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,###,###'

# 틀고정: B2 기준 (첫 행·첫 열 고정)
ws.freeze_panes = 'B2'

# 저장
wb.save(EXCEL_FILE)
print(f"재저장 완료: {EXCEL_FILE}")
