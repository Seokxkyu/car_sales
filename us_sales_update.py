import os
import time
import random
import logging
from datetime import datetime

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import pandas as pd

# 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
EXCEL_FILE = os.path.join(DATA_DIR, 'us_sales_update.xlsx')
SHEET_NAME = 'Brands'
CURRENT_YEAR = datetime.now().year
MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)


def fetch_us_sales():
    """GOODCARBADCAR에서 US 자동차 월별 판매량 스크래핑 반환"""
    HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/115.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://www.goodcarbadcar.net/"
    }
    url = f"https://www.goodcarbadcar.net/{CURRENT_YEAR}-us-auto-sales-figures-by-brand-brand-rankings/"
    session = requests.Session()
    session.headers.update(HEADERS)
    retries = Retry(total=5, backoff_factor=1,
                    status_forcelist=[429,500,502,503,504],
                    allowed_methods=["GET"])
    session.mount("https://", HTTPAdapter(max_retries=retries))
    session.mount("http://", HTTPAdapter(max_retries=retries))

    # 랜덤 딜레이
    time.sleep(random.uniform(1.0, 3.0))
    resp = session.get(url, timeout=10)
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "lxml")
    table = soup.find("table", id="table_6")
    if table is None:
        raise RuntimeError("🚨 could not find US sales table")

    rows = table.find_all("tr", attrs={"data-row-index": True})
    data = [[td.get_text(strip=True).replace(",", "") for td in tr.find_all("td")] for tr in rows]
    df = pd.DataFrame(data, columns=["Brand"] + MONTH_ABBR)
    for m in MONTH_ABBR:
        df[m] = df[m].astype(int)
    return df.set_index('Brand')


def update_us_sales():
    """openpyxl로 기존 시트 손상 없이 CURRENT_YEAR 전체 월 데이터를 삭제 후 덮어쓰기"""
    os.makedirs(DATA_DIR, exist_ok=True)
    sales_df = fetch_us_sales()

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # 브랜드-행 매핑 (Brand은 열 B)
    brand_rows = {ws.cell(row=r, column=2).value: r for r in range(2, ws.max_row+1) if ws.cell(row=r, column=2).value}

    # Jan~Dec 순서로 처리: 기존 컬럼 삭제 후 새로 추가
    for idx, month_abbr in enumerate(MONTH_ABBR, start=1):
        header = f"{CURRENT_YEAR}-{idx:02d}"
        logging.info(f"Updating month '{month_abbr}' as header '{header}'")
        # 기존 컬럼 삭제
        delete_col = None
        for c in range(3, ws.max_column+1):
            if ws.cell(row=1, column=c).value == header:
                delete_col = c
                break
        if delete_col:
            ws.delete_cols(delete_col)
            logging.info(f"Deleted existing column '{header}' at {delete_col}")
        # 새로운 컬럼 추가
        new_col = ws.max_column + 1
        hcell = ws.cell(row=1, column=new_col, value=header)
        hcell.alignment = Alignment(horizontal='center')
        hcell.font = Font(bold=True)
        # 값 작성
        for brand, sales in sales_df[month_abbr].items():
            row = brand_rows.get(brand)
            if not row:
                row = ws.max_row + 1
                ws.cell(row=row, column=2, value=brand)
                brand_rows[brand] = row
            vcell = ws.cell(row=row, column=new_col, value=sales)
            vcell.number_format = '#,###,###'

    # 틀 고정 (C2)
    ws.freeze_panes = 'C2'
    wb.save(EXCEL_FILE)
    logging.info(f"✅ '{EXCEL_FILE}' 업데이트 완료")


if __name__ == '__main__':
    update_us_sales()